import openpyxl as pyxl
from copy import copy

workbook = pyxl.load_workbook("Utv v4.xlsx")
sheet = workbook.active
larare = []

for row in sheet.iter_rows(min_row=1, max_row=1,values_only=True):

    header = row

for row in sheet.iter_rows(min_row=2):
    name = str(row[1].value)
    if name not in larare:
        larare.append(name)
        wb_new = pyxl.Workbook()
        filename = "Utv v4 "+name+".xlsx"
        wb_new.save(filename)

        wb = pyxl.load_workbook(filename)
        ws = wb.active
        ws.append(header)
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.col_idx, value = cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
                print(new_cell._style)

        wb.save(filename)
    else:
        filename = "Utv v4 "+name+".xlsx"
        wb2 = pyxl.load_workbook(filename)
        ws2 = wb2.active
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.col_idx, value = cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)

        wb.save(filename)

